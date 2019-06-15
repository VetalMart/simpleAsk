import os
from sys import exit
from pprint import pprint
from datetime import datetime
from openpyxl import Workbook, load_workbook
from easygui import diropenbox


#ссылка базовой папки
BASE_DIR = diropenbox()
#список файлов в указаной папке
folder_files_names = os.listdir(BASE_DIR)

#список подходящих файлов
xlsx_files_list = []
#список остальных файлов
not_xlsx_files_list = []
#список файлов и листов в них которые - не наш формат заявки
non_format_sheets = []
#список с данными с файлов
datalist = {}
#файл с вкладками заявок
good_data_wb = Workbook()
#файл с вкладками другими
bad_data_wb = Workbook()
#ввести номер заявки
ask_number = int(input("Введи номер последней заявки > "))

def tuple_unzip(t):
    """
    Функция для распаковки екселевских тьюплов
    tuple format: ((value), )
    возвращает список с данными подобный тьюплу
    """
    l = []
    for i in t[0]:
        l.append(i.value)

    return l

def create_range(pos1, pos2, raw):
    """
    Превращает координаты ячеек екселя, в список имен ячеек.
    Нужна для указания места, куда ложить данные. 
    Принимает две буквенные координаты, и номер строки.
    pos1 - первая координата строки
    pos2 - вторая координата строки, может состоять из 2 букв
    raw - номер строки
    """

    if len(pos2) == 1:
        #если одна буква  
        ex_range = ['{0}{1}'.format(chr(i), raw) for i in range(
                        ord(pos1), ord(pos2)+1)]
        #если 2 буквы
    else:
        ex_range = ['{0}{1}'.format(chr(i), raw) for i in range(
                            ord(pos2[0]), 91)]
        double_range = ['{0}{1}{2}'.format('A', chr(i), raw) for i in range(
                            65, ord(pos2[1])+1)]
        ex_range.extend(double_range)

    return ex_range

        
for i in folder_files_names:
    if i.endswith('.xlsx'):
        xlsx_files_list.append(i)
    else:
        not_xlsx_files_list.append(i)
#проверяем не пуст ли список екселевских файлов
if xlsx_files_list:
    #проверяем екселевские файлы - заявки ли?
    for i in xlsx_files_list:
        #загружаем книги по очереди
        wb_temp = load_workbook(filename = os.path.join(BASE_DIR, i))
        #проверяем подходящие листи, проходим по листам в файле
        for j in wb_temp.sheetnames:
            if wb_temp[j]['A1'].value == "!!!":
                #номер с которого начнем
                ask_number += 1
                #если лист подходит, копируем информацию
                #строка
                raw_info = tuple_unzip(wb_temp[j]["A11":"AN11"])
                #фирма
                organization = wb_temp[j]["C12"].value
                #мыло
                email = wb_temp[j]["I12"].value
                #телефон
                phone = wb_temp[j]["B15"].value
                #должность
                occupation = wb_temp[j]["J17"].value
                #имя должостного лица
                name = wb_temp[j]["J16"].value

                # добавляем в главный список
                datalist[ask_number] = {
                    "raw_info": {
                        "date": raw_info[0],
                        "address_region": raw_info[1],
                        "address_city": raw_info[2],
                        "address_street": raw_info[3],
                        "address_build": raw_info[4],
                        "address_flat": raw_info[5],
                        "client": raw_info[6],
                        "mont_org": raw_info[7],
                        "tu": raw_info[8],
                        "project": raw_info[9],
                        "project_org": raw_info[10],
                        "aim": raw_info[11],
                        "gas_pipeline":{
                            "h_pressure": {
                                "rozp": {
                                    "d": {
                                        "pe": raw_info[12],
                                        "steel": raw_info[13],
                                    },
                                    "l": raw_info[14],
                                },
                                "vvid": {
                                    "d": {
                                        "pe": raw_info[15],
                                        "steel": raw_info[16],
                                    },
                                    "l": raw_info[17],
                                }
                            },
                            "l_pressure": {
                                "rozp": {
                                    "d": {
                                        "pe": raw_info[18],
                                        "steel": raw_info[19],
                                    },
                                    "l": raw_info[20],
                                },
                                "vvid": {
                                    "d": {
                                        "pe": raw_info[21],
                                        "steel": raw_info[22],
                                    },
                                    "l": raw_info[23],
                                },
                                "vvidnyi": {
                                    "d": raw_info[24],
                                    "l": raw_info[25],
                                }
                            }
                        },
                        "regulator": {
                            "GRP": {
                                "type": raw_info[26],
                                "amount": raw_info[27],
                            },
                            "RT": {
                                "type": raw_info[28],
                                "amount": raw_info[29],
                            }
                        },
                        "gso": {
                            "aim": raw_info[30],
                            "info": raw_info[31],
                            "gas_consume": raw_info[32],
                        },
                        "vog": {
                            "type": raw_info[33],
                            "model": raw_info[34],
                            "amount": raw_info[35],
                        }, 
                        "korretor": {
                            "model": raw_info[36],
                            "amount": raw_info[37],
                        },
                        "building": {
                            "flat": raw_info[38],
                            "level": raw_info[39],
                        },
                    },
                    "organization": organization,
                    "email": email,
                    "phone": phone,
                    "occupation": occupation,
                    "name": name,
                }       
            else:
                non_format_sheets.append('файл: {0}, лист: {1}'.format(i, j))
#выходим с программы
else:
    print('В папке отсутствуют файлы ексель.')
    sys.exit()


#открываем шаблон для заявок
good_data_wb = load_workbook(filename = os.path.join(
                BASE_DIR, 'template.xlsx'))

#открываем шаблон журнала
magazin_template = load_workbook(filename = os.path.join(
                BASE_DIR, 'magazin.xlsx'))

ws_ask = good_data_wb['л1']
ws_mag = magazin_template['к']
rnm = 7     #строка с которой начнем счет в журнале (-1)
#цикл который наполняет екселевские файлы
for k, v in datalist.items():
    #создаем копию листа в файле с заявками
    new_ws = good_data_wb.copy_worksheet(ws_ask)
    #переименовываем его под номер заявки  
    new_ws.title = str(k)
    new_ws['C11'] = v['organization']
    new_ws.merge_cells('C11:G11')
    new_ws['B14'] = v['phone']
    new_ws['I11'] = v['email']
    new_ws.merge_cells('I11:L11')
    new_ws['J15'] = v['name']
    new_ws.merge_cells('J15:L15')
    new_ws['J16'] = v['occupation']
    new_ws.merge_cells('J16:L16')
    new_ws['L13'] = str(k)
    new_ws['L14'] = datetime.today().strftime("%d.%m.%Y") 
    new_ws['A10'] = v['raw_info']['date']
    new_ws['B10'] = v['raw_info']['address_region']
    new_ws['C10'] = v['raw_info']['address_city']
    new_ws['D10'] = v['raw_info']['address_street']
    new_ws['E10'] = v['raw_info']['address_build']
    new_ws['F10'] = v['raw_info']['address_flat']
    new_ws['G10'] = v['raw_info']['client']
    new_ws['H10'] = v['raw_info']['mont_org']
    new_ws['I10'] = v['raw_info']['tu']
    new_ws['J10'] = v['raw_info']['project']
    new_ws['K10'] = v['raw_info']['project_org']
    new_ws['L10'] = v['raw_info']['aim']
    new_ws['M10'] = v['raw_info']['gas_pipeline']['h_pressure']['rozp']['d']['pe']
    new_ws['N10'] = v['raw_info']['gas_pipeline']['h_pressure']['rozp']['d']['steel']
    new_ws['O10'] = v['raw_info']['gas_pipeline']['h_pressure']['rozp']['l']
    new_ws['P10'] = v['raw_info']['gas_pipeline']['h_pressure']['vvid']['d']['pe']
    new_ws['Q10'] = v['raw_info']['gas_pipeline']['h_pressure']['vvid']['d']['steel']
    new_ws['R10'] = v['raw_info']['gas_pipeline']['h_pressure']['vvid']['l']
    new_ws['S10'] = v['raw_info']['gas_pipeline']['l_pressure']['rozp']['d']['pe']
    new_ws['T10'] = v['raw_info']['gas_pipeline']['l_pressure']['rozp']['d']['steel']
    new_ws['U10'] = v['raw_info']['gas_pipeline']['l_pressure']['rozp']['l']
    new_ws['V10'] = v['raw_info']['gas_pipeline']['l_pressure']['vvid']['d']['pe']
    new_ws['W10'] = v['raw_info']['gas_pipeline']['l_pressure']['vvid']['d']['steel']
    new_ws['X10'] = v['raw_info']['gas_pipeline']['l_pressure']['vvid']['l']
    new_ws['Y10'] = v['raw_info']['gas_pipeline']['l_pressure']['vvidnyi']['d']
    new_ws['Z10'] = v['raw_info']['gas_pipeline']['l_pressure']['vvidnyi']['l']
    new_ws['AA10'] = v['raw_info']['regulator']['GRP']['type']
    new_ws['AB10'] = v['raw_info']['regulator']['GRP']['amount']
    new_ws['AC10'] = v['raw_info']['regulator']['RT']['type']
    new_ws['AD10'] = v['raw_info']['regulator']['RT']['amount']
    new_ws['AE10'] = v['raw_info']['gso']['aim']
    new_ws['AF10'] = v['raw_info']['gso']['info']
    new_ws['AG10'] = v['raw_info']['gso']['gas_consume']
    new_ws['AH10'] = v['raw_info']['vog']['type']
    new_ws['AI10'] = v['raw_info']['vog']['model']
    new_ws['AJ10'] = v['raw_info']['vog']['amount']
    new_ws['AK10'] = v['raw_info']['korretor']['model']
    new_ws['AL10'] = v['raw_info']['korretor']['amount']
    new_ws['AM10'] = v['raw_info']['building']['flat']
    new_ws['AN10'] = v['raw_info']['building']['level']

    new_ws.print_area = 'A1:AN16'

    #теперь работаем с прототипом журнала
    rnm += 1
    ws_mag['A{0}'.format(rnm)] = str(k)
    ws_mag['I{0}'.format(rnm)] = str(k)
    ws_mag['J{0}'.format(rnm)] = datetime.today().strftime("%d.%m.%Y") 
    ws_mag['C{0}'.format(rnm)] = v['raw_info']['address_region']
    ws_mag['D{0}'.format(rnm)] = v['raw_info']['address_city']
    ws_mag['E{0}'.format(rnm)] = v['raw_info']['address_street']
    ws_mag['F{0}'.format(rnm)] = v['raw_info']['address_build']
    ws_mag['G{0}'.format(rnm)] = v['raw_info']['address_flat']
    ws_mag['X{0}'.format(rnm)] = v['raw_info']['client']
    ws_mag['H{0}'.format(rnm)] = v['raw_info']['mont_org']
    ws_mag['U{0}'.format(rnm)] = v['raw_info']['tu']
    ws_mag['V{0}'.format(rnm)] = v['raw_info']['project']
    ws_mag['W{0}'.format(rnm)] = v['raw_info']['project_org']
    ws_mag['AA{0}'.format(rnm)] = v['raw_info']['aim']
    ws_mag['AB{0}'.format(rnm)] = v['raw_info']['gas_pipeline']['h_pressure']['rozp']['d']['pe']
    ws_mag['AC{0}'.format(rnm)] = v['raw_info']['gas_pipeline']['h_pressure']['rozp']['d']['steel']
    ws_mag['AD{0}'.format(rnm)] = v['raw_info']['gas_pipeline']['h_pressure']['rozp']['l']
    ws_mag['AE{0}'.format(rnm)] = v['raw_info']['gas_pipeline']['h_pressure']['vvid']['d']['pe']
    ws_mag['AF{0}'.format(rnm)] = v['raw_info']['gas_pipeline']['h_pressure']['vvid']['d']['steel']
    ws_mag['AG{0}'.format(rnm)] = v['raw_info']['gas_pipeline']['h_pressure']['vvid']['l']
    ws_mag['AH{0}'.format(rnm)] = v['raw_info']['gas_pipeline']['l_pressure']['rozp']['d']['pe']
    ws_mag['AI{0}'.format(rnm)] = v['raw_info']['gas_pipeline']['l_pressure']['rozp']['d']['steel']
    ws_mag['AJ{0}'.format(rnm)] = v['raw_info']['gas_pipeline']['l_pressure']['rozp']['l']
    ws_mag['AK{0}'.format(rnm)] = v['raw_info']['gas_pipeline']['l_pressure']['vvid']['d']['pe']
    ws_mag['AL{0}'.format(rnm)] = v['raw_info']['gas_pipeline']['l_pressure']['vvid']['d']['steel']
    ws_mag['AM{0}'.format(rnm)] = v['raw_info']['gas_pipeline']['l_pressure']['vvid']['l']
    ws_mag['AN{0}'.format(rnm)] = v['raw_info']['gas_pipeline']['l_pressure']['vvidnyi']['d']
    ws_mag['AO{0}'.format(rnm)] = v['raw_info']['gas_pipeline']['l_pressure']['vvidnyi']['l']
    ws_mag['AP{0}'.format(rnm)] = v['raw_info']['regulator']['GRP']['type']
    ws_mag['AQ{0}'.format(rnm)] = v['raw_info']['regulator']['GRP']['amount']
    ws_mag['AR{0}'.format(rnm)] = v['raw_info']['regulator']['RT']['type']
    ws_mag['AS{0}'.format(rnm)] = v['raw_info']['regulator']['RT']['amount']
    ws_mag['AT{0}'.format(rnm)] = v['raw_info']['gso']['aim']
    ws_mag['AU{0}'.format(rnm)] = v['raw_info']['gso']['info']
    ws_mag['AV{0}'.format(rnm)] = v['raw_info']['gso']['gas_consume']
    ws_mag['AW{0}'.format(rnm)] = v['raw_info']['vog']['type']
    ws_mag['AX{0}'.format(rnm)] = v['raw_info']['vog']['model']
    ws_mag['AY{0}'.format(rnm)] = v['raw_info']['vog']['amount']
    ws_mag['AZ{0}'.format(rnm)] = v['raw_info']['korretor']['model']
    ws_mag['BA{0}'.format(rnm)] = v['raw_info']['korretor']['amount']
    ws_mag['BB{0}'.format(rnm)] = v['raw_info']['building']['flat']
    ws_mag['BC{0}'.format(rnm)] = v['raw_info']['building']['level']

#сохраняем файлы
good_data_wb.save('файл_с_заявками.xlsx') 
magazin_template.save('шаблон_журнала_с_заявками.xlsx') 

#файлы и листы которые "не вошли"
non_satisfy = [not_xlsx_files_list, non_format_sheets]
with open('отбросы.txt', 'w') as file:
    print("Это список файлов в этой папке, которые \
        не *.xlsx\n", file=file, sep='\n')
    print(*not_xlsx_files_list, file=file, sep='\n')
    print("\n\nЭто список файлов и листов в них, которые не\
        соответствуют критерию !!!\n", file=file, sep='\n')
    print(*non_format_sheets, file=file, sep='\n')
